import json
import csv
import os
import openpyxl
import time
import urllib.request
import re
from datetime import datetime

class TeamsNotifier:
    def __init__(self, webhook_url, enabled=False):
        self.webhook_url = webhook_url
        self.enabled = enabled
        self.template_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "templates", "teams_notification.json")

    def parse_exception_name(self, raw_name):
        """Advanced parser to extract clean human-readable messages from messy Dynatrace logs."""
        if not raw_name:
            return "Unknown Exception"
            
        # 1. Eager whitespace cleaning
        clean = str(raw_name).replace("EMSG:", "")
        clean = clean.replace('\r', ' ').replace('\n', ' ').replace('\t', ' ')
        clean = re.sub(r'\s+', ' ', clean).strip()
        
        # 2. Tag Stitching & Fragment Merging
        # Often logs look like: "shows that the XXX STRACE:the XXX NISSAN..."
        # We want to remove the tag but keep the text if it's non-technical human text.
        
        # Define technical markers
        tech_markers = r'(?:com\.|gw\.|deployment\.|java\.|org\.|edge\.|android\.|apple\.)'
        
        # Rule: If a tag (STRACE:) is followed by human text (not a package), replace tag with space
        # Pattern looks for a tag followed by something that IS NOT a package marker
        def stitch_tags(text):
            # Split into chunks by tags
            parts = re.split(r'(STRACE:|TRACE:|STRACERAW:)', text, flags=re.IGNORECASE)
            if len(parts) <= 1:
                return text
                
            result = parts[0]
            for i in range(1, len(parts), 2):
                tag = parts[i]
                next_text = parts[i+1].strip() if i+1 < len(parts) else ""
                
                # If next text starts with a package, cut everything from here
                if re.match(rf'^{tech_markers}', next_text, re.IGNORECASE):
                    break
                
                # Otherwise, replace tag with space and continue stitching
                result += " " + next_text
            return result.strip()
            
        clean = stitch_tags(clean)
        
        # 3. Deduplication of fragments
        # Fixes cases like "the XXX the XXX" or "is effective on the date: XXX is effective on the date: XXX"
        def deduplicate_fragments(text):
            # Normalization helper
            def normalize(s):
                return re.sub(r'[^a-zA-Z0-9]', '', s).lower()

            # Split by common Dynatrace placeholders/separators
            parts = [p.strip() for p in text.split('XXX')]
            unique_parts = []
            seen_normalized = set()
            
            for p in parts:
                if not p:
                    continue
                norm = normalize(p)
                # If this fragment (when stripped of symbols) is seen or is a substring of what we have
                is_rep = False
                for existing in seen_normalized:
                    if norm in existing or existing in norm:
                        if len(norm) >= 2: # Keep dedupe threshold low for "the", "is", etc.
                            is_rep = True
                            break
                
                if not is_rep:
                    unique_parts.append(p)
                    seen_normalized.add(norm)
            
            return " XXX ".join(unique_parts)

        clean = deduplicate_fragments(clean)

        # 4. Handle 'at ' split safely (ONLY on package boundaries)
        trace_match = re.search(rf'\bat\s+{tech_markers}', clean, re.IGNORECASE)
        if trace_match:
            # Only cut if we have a decent amount of message already, or if it's very clearly a trace
            if trace_match.start() > 10:
                clean = clean[:trace_match.start()].strip()

        # 5. Quote Balancing
        # If we have an odd number of quotes, try to see if the next part of the raw string finishes it
        if clean.count('"') % 2 != 0:
            # Simplest fix: just remove trailing unclosed quotes for the alert
            clean = clean.strip('"')

        # Final Cleanup
        clean = clean.strip(' ,:;-')
        clean = re.sub(r'\s+', ' ', clean).strip()
        
        # Truncate only if massive
        if len(clean) > 800:
            clean = clean[:797] + "..."
            
        return clean if clean else "Technical Exception (See Excel for details)"

    def get_business_summary(self, clean_name, raw_name):
        """Translates technical/messy logs into an ultra-concise business summary."""
        mappings = [
            (r"vehicle history activity", "Vehicle History Flagged"),
            (r"stringconstraintshandlerbase", "Input Value Too Long"),
            (r"validatestringlength", "Input Value Too Long"),
            (r"license already present", "Agent License Conflict"),
            (r"could not find a policy", "Policy Search Failed"),
            (r"database bean version conflict", "System Sync Error (Retry)"),
            (r"concurrentdatachangeexception", "System Sync Error (Retry)"),
            (r"locked branch", "Record Locked"),
            (r"no underwriting company", "Underwriting Config Missing"),
            (r"unable to parse xml", "Data Format Error"),
            (r"lhpi ordering failed", "Third-Party System Down"),
            (r"ineligible for farmers auto", "Ineligible for Auto Policy"),
            (r"expiration date cannot be before", "Effective Date Error"),
            (r"validationerror", "Data Validation Error")
        ]
        
        search_text = (str(clean_name) + " " + str(raw_name)).lower()
        
        # Specific KT-based technical mappings
        if "adrin miley" in search_text:
            return "Known Technical Context (Adrin Miley)"
        
        for pattern, business_phrase in mappings:
            if re.search(pattern, search_text):
                return business_phrase
        
        return "Business Validation Alert"

    def send_alert(self, exception_name, count, baseline, status, category="N/A"):
        if not self.enabled or not self.webhook_url or self.webhook_url == "YOUR_WEBHOOK_URL_HERE":
            print(f"[SIMULATED TEAMS ALERT] Status: {status} | Category: {category} | Exception: {exception_name}")
            return

        try:
            # Parse and clean the exception name for the notification card
            clean_name = self.parse_exception_name(exception_name)
            business_summary = self.get_business_summary(clean_name, exception_name)

            with open(self.template_path, 'r') as f:
                card_data = json.load(f)

            # Dynamic values
            status_color = "Attention" if status == "New" else "Warning"
            action_taken = "Immediate Alert" if status == "New" else "Rising Trend Confirmed"
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Helper to recursively replace placeholders in the dictionary
            def replace_placeholders(obj):
                if isinstance(obj, str):
                    obj = obj.replace("${status}", status)
                    obj = obj.replace("${statusColor}", status_color)
                    obj = obj.replace("${category}", category)
                    # We make the Business Summary the focal point
                    obj = obj.replace("${exceptionName}", f"**{business_summary}**\n\n_{clean_name}_")
                    obj = obj.replace("${currentCount}", str(count))
                    obj = obj.replace("${baselineCount}", str(baseline))
                    obj = obj.replace("${timestamp}", timestamp)
                    obj = obj.replace("${actionTaken}", action_taken)
                    return obj
                elif isinstance(obj, list):
                    return [replace_placeholders(item) for item in obj]
                elif isinstance(obj, dict):
                    return {k: replace_placeholders(v) for k, v in obj.items()}
                return obj

            card_data = replace_placeholders(card_data)
            
            message_payload = {
                "type": "message",
                "attachments": [
                    {
                        "contentType": "application/vnd.microsoft.card.adaptive",
                        "content": card_data
                    }
                ]
            }

            req = urllib.request.Request(
                self.webhook_url,
                data=json.dumps(message_payload).encode('utf-8'),
                headers={'Content-Type': 'application/json'}
            )
            with urllib.request.urlopen(req) as response:
                status_code = response.getcode()
                resp_data = response.read().decode('utf-8')
                if status_code in [200, 202]:
                    print(f"✅ Teams Alert Sent: {business_summary}")
                else:
                    print(f"❌ Failed to send Teams alert. Status: {status_code} | Response: {resp_data}")
        except Exception as e:
            print(f"Error sending Teams notification: {e}")

def load_config():
    config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config.json")
    if not os.path.exists(config_path):
        return {}
    with open(config_path, 'r') as f:
        return json.load(f)

def get_user_input(prompt):
    """Abstraction for user input to allow GUI overrides."""
    return input(prompt)

def normalize_exception_name(name):
    """Strips leading/trailing quotes and redundant whitespace for consistent matching."""
    if not name:
        return ""
    # Remove leading/trailing quotes (single and double)
    clean = str(name).strip().strip("'").strip('"')
    # Collapse multiple spaces
    clean = re.sub(r'\s+', ' ', clean).strip()
    return clean

def load_real_input(filepath):
    # Use a dictionary to aggregate counts by normalized name
    aggregated = {}
    if not os.path.exists(filepath):
        print(f"Error: Input file {filepath} not found.")
        return []
    
    with open(filepath, mode='r', encoding='utf-8', newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            raw_name = row.get('exception', '').strip()
            if not raw_name:
                continue
                
            count_raw = row.get('Count', '0')
            count = int(count_raw) if str(count_raw).isdigit() else 0
            
            norm_name = normalize_exception_name(raw_name)
            if norm_name in aggregated:
                aggregated[norm_name]['count'] += count
                # We keep the first raw_name encountered for display
            else:
                aggregated[norm_name] = {
                    'name': norm_name,
                    'count': count,
                    'raw_name': raw_name
                }
    
    return list(aggregated.values())

def update_excel_workbook(workbook_path, current_data):
    if not os.path.exists(workbook_path):
        print(f"Error: Workbook {workbook_path} not found.")
        return
    
    wb = openpyxl.load_workbook(workbook_path)
    if 'Current' not in wb.sheetnames:
        print("Error: 'Current' sheet not found in workbook.")
        return
    
    sheet = wb['Current']
    max_row = sheet.max_row
    if max_row > 1:
        for row in sheet.iter_rows(min_row=2, max_row=max_row):
            for cell in row:
                cell.value = None

    for i, entry in enumerate(current_data, start=2):
        sheet.cell(row=i, column=1).value = entry['name']
        sheet.cell(row=i, column=2).value = entry['count']
        sheet.cell(row=i, column=3).value = f"=IFERROR(VLOOKUP(A{i},'30Days-Baseline'!A:D,2,FALSE),0)"
        sheet.cell(row=i, column=4).value = f"=IFERROR(IF(B{i}>ROUNDUP(C{i},0),\"Elevated\",\"Ok\"),\"New\")"
        sheet.cell(row=i, column=5).value = f"=IFERROR(VLOOKUP(A{i},'30Days-Baseline'!A:E,5,FALSE),\"\")"

    wb.save(workbook_path)

def load_baseline_from_excel(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if '30Days-Baseline' not in wb.sheetnames:
        return {}
    
    sheet = wb['30Days-Baseline']
    baseline = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        normalized_name = normalize_exception_name(row[0])
        baseline[normalized_name] = row[1] if isinstance(row[1], (int, float)) else 0
    return baseline

def merge_new_exceptions_to_baseline(workbook_path, new_exceptions):
    """Appends newly discovered exceptions to the baseline sheet with a count of 0."""
    if not new_exceptions:
        return
        
    try:
        wb = openpyxl.load_workbook(workbook_path)
        if '30Days-Baseline' not in wb.sheetnames:
            print("Error: '30Days-Baseline' sheet not found for merging.")
            return
            
        sheet = wb['30Days-Baseline']
        current_max = sheet.max_row
        
        # Build set of existing normalized names to avoid duplicates
        existing_names = {normalize_exception_name(row[0].value) for row in sheet.iter_rows(min_row=2, max_col=1) if row[0].value}
        
        added_count = 0
        for exc in new_exceptions:
            norm_name = exc['name']
            if norm_name not in existing_names:
                sheet.cell(row=current_max + 1 + added_count, column=1).value = exc['raw_name']
                sheet.cell(row=current_max + 1 + added_count, column=2).value = 0
                existing_names.add(norm_name)
                added_count += 1
        
        if added_count > 0:
            wb.save(workbook_path)
            print(f"✅ Baseline Updated: Added {added_count} new exceptions to '30Days-Baseline'.")
            
    except Exception as e:
        print(f"Error updating baseline with new exceptions: {e}")

def monitor_cycle(notifier, config, project_root, input_file=None):
    data_path = os.path.join(project_root, config.get("DATA_DIR", "data"))
    if input_file is None:
        input_csv = os.path.join(data_path, config.get("INPUT_CSV", "table-data (2).csv"))
    else:
        input_csv = os.path.join(data_path, input_file)
        
    workbook_file = os.path.join(data_path, config.get("BASELINE_EXCEL", "PC_Prod_Exception_Monitoring_R02_2026-DT 2.xlsx"))
    
    print(f"\n[{datetime.now().strftime('%H:%M:%S')}] Starting Monitoring Cycle (Input: {os.path.basename(input_csv)})...")
    
    current_data = load_real_input(input_csv)
    baseline = load_baseline_from_excel(workbook_file)
    update_excel_workbook(workbook_file, current_data)
    
    findings = {"New": [], "Elevated": [], "Ok": []}
    for entry in current_data:
        name, count = entry['name'], entry['count']
        raw_name = entry['raw_name']
        base_count = baseline.get(name)
        
        if base_count is None:
            findings["New"].append({"name": name, "count": count, "baseline": 0, "raw_name": raw_name})
        elif count > base_count:
            findings["Elevated"].append({"name": name, "count": count, "baseline": base_count, "raw_name": raw_name})
        else:
            findings["Ok"].append({"name": name, "count": count, "baseline": base_count, "raw_name": raw_name})
            
    return findings

def main():
    config = load_config()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    notifier = TeamsNotifier(config.get("TEAMS_WEBHOOK_URL"), config.get("ENABLE_NOTIFICATIONS", False))
    
    print("--- Unified Exception Monitoring Automation ---")
    
    # 1. Initial Processing (Cycle 1)
    findings = monitor_cycle(notifier, config, project_root)
    
    # 2. Handle NEW Exceptions Immediately
    # Note: Elevated exceptions are held for Cycle 2 verification
    for exc in findings["New"]:
        notifier.send_alert(exc["raw_name"], exc["count"], exc["baseline"], "New", category="New")
    
    # 3. Handle ELEVATED or NEW Verification (Cycle 2)
    if findings["Elevated"] or findings["New"]:
        print(f"\nDetected {len(findings['Elevated'])} ELEVATED and {len(findings['New'])} NEW exceptions.")
        
        # Check for Manual Mode
        if config.get("MANUAL_MODE", False):
            print("\n[MANUAL MODE ACTIVE]")
            print(f"Cycle 1 is complete. Please download the second Dynatrace report (Cycle 2) and save it as: {config.get('SECOND_INPUT_CSV')}")
            get_user_input("Press Enter once the file is ready to continue Cycle 2...")
        else:
            wait_time = config.get("DELAY_MINUTES", 30)
            print(f"Waiting {wait_time} minute(s) to verify trend and check for updates...")
            time.sleep(wait_time * 60)
        
        try:
            print("\nPerforming Second Monitoring Cycle...")
            second_input = config.get("SECOND_INPUT_CSV", config.get("INPUT_CSV"))
            second_findings = monitor_cycle(notifier, config, project_root, input_file=second_input)
            
            # Lookup for first findings
            first_counts = {e['name']: e['count'] for e in (findings["Elevated"] + findings["New"])}
            first_new_names = {e['name'] for e in findings["New"]}
            
            # Handle Cycle 2 Findings
            # Combine New and Elevated for comprehensive trend checking
            all_second_findings = second_findings["New"] + second_findings["Elevated"]
            
            # Keep track of what we've alerted to avoid duplicates
            alerted_names = set()

            for exc in all_second_findings:
                name = exc['name']
                count = exc['count']
                raw_name = exc['raw_name']
                
                # If it's Elevated in Cycle 2, it's generally worth an alert because:
                # 1. If it was New in Cycle 1, we check if it's still rising (already alerted once, so we only alert if it gained)
                # 2. If it was Elevated in Cycle 1, we haven't alerted it yet! So we MUST alert it now as long as it's still Elevated.
                
                if name in first_counts:
                    # Case A: Elevated in Cycle 1
                    if name not in first_new_names:
                        # We didn't alert this in Cycle 1, so alert it now as "Confirmed High"
                        # even if it didn't increase from Cycle 1 (as long as it's still > Baseline)
                        print(f"High trend confirmed for: {name[:50]} (Count: {count})")
                        notifier.send_alert(raw_name, count, exc["baseline"], "Elevated", category="Elevated")
                        alerted_names.add(name)
                    
                    # Case B: New in Cycle 1
                    elif count > first_counts[name]:
                        # It was New in Cycle 1 and it's INCREASING further
                        print(f"Rising trend confirmed for New exception: {name[:50]} ({first_counts[name]} -> {count})")
                        notifier.send_alert(raw_name, count, exc["baseline"], "New (Still Increasing)", category="Elevated")
                        alerted_names.add(name)
                    else:
                        print(f"Stabilized for New exception: {name[:50]} (Count: {count})")
                else:
                    # Case C: Completely fresh in Cycle 2
                    print(f"Update: Fresh exception detected in Cycle 2: {name[:50]}")
                    status = "New" if exc in second_findings["New"] else "Elevated"
                    category = "New" if status == "New" else "Elevated"
                    notifier.send_alert(raw_name, count, exc["baseline"], status, category=category)
                    alerted_names.add(name)
                    
        except KeyboardInterrupt:
            print("\nMonitoring interrupted by user.")

    # 4. Update Baseline with NEW exceptions for next run
    if findings["New"]:
        workbook_file = os.path.join(project_root, config.get("DATA_DIR", "data"), config.get("BASELINE_EXCEL"))
        merge_new_exceptions_to_baseline(workbook_file, findings["New"])

    print("\nProcessing Complete.")

if __name__ == "__main__":
    main()
